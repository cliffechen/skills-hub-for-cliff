#!/usr/bin/env python
"""Parse a Sif keyword Excel workbook into reusable JSON/Markdown summaries."""

from __future__ import annotations

import argparse
import csv
import json
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


CLUSTER_RULES: list[tuple[str, list[str]]] = [
    ("c15_core", ["c15", "c15:0", "pentadecanoic", "fatty15", "fatty 15"]),
    ("nmn_nad", ["nmn", "nad+", "nad supplement", "nicotinamide mononucleotide"]),
    ("pqq", ["pqq", "pyrroloquinoline"]),
    ("coq10", ["coq10", "co q10", "coenzyme q10", "ubiquinol", "ubiquinone"]),
    ("urolithin", ["urolithin", "mitopure"]),
    ("phosphatidylcholine", ["phosphatidylcholine"]),
    ("l_carnitine", ["carnitine"]),
    ("tocotrienol", ["tocotrienol"]),
    ("longevity_mito", ["longevity", "mitochondria", "mitochondrial", "anti aging", "cellular energy"]),
    ("omega_fatty", ["omega", "fatty acid"]),
    ("vitamin_c_noise", ["vitamin c", "liposomal vitamin c", "ascorbic"]),
    ("risk_noise", ["bpc", "delta 9", "thc", "hgh", "peptide", "semaglutide", "steroid"]),
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("excel", help="Path to Sif Excel workbook.")
    parser.add_argument("--output-dir", required=True, help="Directory for summary outputs.")
    parser.add_argument("--name", default="sif-keyword", help="Output file prefix.")
    parser.add_argument("--top-n", type=int, default=30, help="Number of top keywords to export.")
    return parser.parse_args()


def normalize_header(value: Any) -> str:
    return str(value or "").strip()


def to_number(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(",", "").replace("%", "").strip()
    try:
        return float(text)
    except ValueError:
        return 0.0


def detect_cluster(keyword: str) -> str:
    low = keyword.lower()
    for cluster, terms in CLUSTER_RULES:
        if any(term in low for term in terms):
            return cluster
    return "other"


def row_to_dict(headers: list[str], row: tuple[Any, ...]) -> dict[str, Any]:
    return {headers[index]: row[index] if index < len(row) else None for index in range(len(headers))}


def get_value(row: dict[str, Any], candidates: list[str]) -> Any:
    for candidate in candidates:
        if candidate in row:
            return row[candidate]
    return None


def main() -> None:
    args = parse_args()
    excel = Path(args.excel).expanduser().resolve()
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(excel, read_only=True, data_only=True)
    rows: list[dict[str, Any]] = []
    sheet_stats = []

    for ws in wb.worksheets:
        iterator = ws.iter_rows(values_only=True)
        try:
            headers = [normalize_header(cell) for cell in next(iterator)]
        except StopIteration:
            continue
        sheet_count = 0
        for raw_row in iterator:
            if not any(cell is not None for cell in raw_row):
                continue
            item = row_to_dict(headers, raw_row)
            keyword = str(get_value(item, ["关键词", "keyword", "Keyword"]) or "").strip()
            relevance = str(get_value(item, ["相关性", "relevance", "Relevance"]) or "").strip()
            weekly_search = to_number(get_value(item, ["周搜索量", "weekly_search", "Weekly Search Volume"]))
            score = to_number(get_value(item, ["相关性得分", "score", "Score"]))
            item["_sheet"] = ws.title
            item["_keyword"] = keyword
            item["_relevance"] = relevance or ws.title
            item["_weekly_search"] = weekly_search
            item["_score"] = score
            item["_cluster"] = detect_cluster(keyword)
            rows.append(item)
            sheet_count += 1
        sheet_stats.append({"sheet": ws.title, "rows": sheet_count})

    relevance_counts = Counter(row["_relevance"] for row in rows)
    cluster_counts = Counter(row["_cluster"] for row in rows)
    weekly_by_relevance: dict[str, float] = defaultdict(float)
    weekly_by_cluster: dict[str, float] = defaultdict(float)
    for row in rows:
        weekly_by_relevance[row["_relevance"]] += row["_weekly_search"]
        weekly_by_cluster[row["_cluster"]] += row["_weekly_search"]

    top_keywords = sorted(rows, key=lambda item: (item["_score"], item["_weekly_search"]), reverse=True)[: args.top_n]
    high_volume_low_relevance = sorted(
        [row for row in rows if row["_weekly_search"] > 0 and re.search(r"低|不相关|low|irrelevant", row["_relevance"], re.I)],
        key=lambda item: item["_weekly_search"],
        reverse=True,
    )[: args.top_n]

    summary = {
        "excel": str(excel),
        "total_rows": len(rows),
        "sheets": sheet_stats,
        "relevance_counts": dict(relevance_counts),
        "weekly_by_relevance": dict(sorted(weekly_by_relevance.items(), key=lambda item: item[1], reverse=True)),
        "cluster_counts": dict(cluster_counts),
        "weekly_by_cluster": dict(sorted(weekly_by_cluster.items(), key=lambda item: item[1], reverse=True)),
        "top_keywords": [
            {
                "keyword": row["_keyword"],
                "relevance": row["_relevance"],
                "score": row["_score"],
                "weekly_search": row["_weekly_search"],
                "cluster": row["_cluster"],
            }
            for row in top_keywords
        ],
        "high_volume_low_relevance": [
            {
                "keyword": row["_keyword"],
                "relevance": row["_relevance"],
                "score": row["_score"],
                "weekly_search": row["_weekly_search"],
                "cluster": row["_cluster"],
            }
            for row in high_volume_low_relevance
        ],
    }

    json_path = output_dir / f"{args.name}-summary.json"
    json_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")

    md_path = output_dir / f"{args.name}-summary.md"
    md_lines = [
        "# Sif Keyword Summary",
        "",
        f"- Source: `{excel}`",
        f"- Total rows: `{len(rows)}`",
        "",
        "## Sheets",
        "",
        "| Sheet | Rows |",
        "| --- | ---: |",
    ]
    md_lines += [f"| {item['sheet']} | {item['rows']} |" for item in sheet_stats]
    md_lines += ["", "## Cluster Weekly Search", "", "| Cluster | Rows | Weekly Search |", "| --- | ---: | ---: |"]
    for cluster, weekly in sorted(weekly_by_cluster.items(), key=lambda item: item[1], reverse=True):
        md_lines.append(f"| {cluster} | {cluster_counts[cluster]} | {weekly:.0f} |")
    md_lines += ["", "## Top Keywords", "", "| Keyword | Relevance | Score | Weekly Search | Cluster |", "| --- | --- | ---: | ---: | --- |"]
    for row in summary["top_keywords"]:
        md_lines.append(f"| {row['keyword']} | {row['relevance']} | {row['score']:.2f} | {row['weekly_search']:.0f} | {row['cluster']} |")
    md_lines += ["", "## High-Volume Low-Relevance Watchlist", "", "| Keyword | Relevance | Score | Weekly Search | Cluster |", "| --- | --- | ---: | ---: | --- |"]
    for row in summary["high_volume_low_relevance"]:
        md_lines.append(f"| {row['keyword']} | {row['relevance']} | {row['score']:.2f} | {row['weekly_search']:.0f} | {row['cluster']} |")
    md_path.write_text("\n".join(md_lines) + "\n", encoding="utf-8")

    csv_path = output_dir / f"{args.name}-top-keywords.csv"
    with csv_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=["keyword", "relevance", "score", "weekly_search", "cluster"])
        writer.writeheader()
        writer.writerows(summary["top_keywords"])

    print(json.dumps({"json": str(json_path), "markdown": str(md_path), "csv": str(csv_path)}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
